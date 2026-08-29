"""Книга Excel: обзор, задачи, лента ячейками, связи, смета, скоркард, разговор.

Даты кладутся настоящими датами, деньги и проценты — числами, итоги сметы —
формулами. Иначе книга превращается в картинку таблицы: получатель не может ни
отсортировать по сроку, ни отфильтровать по статусу, ни поправить ставку и
увидеть новый итог, — то есть не получает ничего, чего не давал бы PDF.
"""

from datetime import date, timedelta
from io import BytesIO

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule, DataBarRule
from openpyxl.styles import Alignment, Border, Font, NamedStyle, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from app.export import budget, theme
from app.export.budget import Zoom
from app.export.document import DocTask, ExportDocument, ExportSection

FONT = theme.XLSX_FONT

#: Рост строки. Задаётся явно: на листе ленты высота строки — это высота
#: полоски, и значение по умолчанию превращает диаграмму в набор нитей.
ROW_H = 18.0

#: Ширина колонки одного дня на листе ленты, в знаках. Меньше — и полоска
#: короткой задачи вырождается в засечку.
DAY_COL_W = 3.0

_thin = Side(style="thin", color=theme.BORDER)


def _styles(wb: Workbook) -> None:
    """Именованные стили — один раз на книгу, а не на ячейку: иначе размер
    файла растёт числом ячеек, а не числом видов ячеек."""

    def add(name: str, **kw) -> None:
        style = NamedStyle(name=name)
        style.font = kw.get("font", Font(name=FONT, size=10, color=theme.TEXT))
        style.alignment = kw.get("align", Alignment(vertical="center"))
        if "fill" in kw:
            style.fill = kw["fill"]
        if kw.get("rule", True):
            style.border = Border(bottom=_thin)
        if "number_format" in kw:
            style.number_format = kw["number_format"]
        wb.add_named_style(style)

    right = Alignment(vertical="center", horizontal="right")

    add("x-h1", font=Font(name=FONT, size=20, bold=True, color=theme.TEXT), rule=False)
    add("x-h2", font=Font(name=FONT, size=13, bold=True, color=theme.TEXT), rule=False)
    add("x-muted", font=Font(name=FONT, size=10, color=theme.TEXT_MUTED), rule=False)
    add(
        "x-head",
        font=Font(name=FONT, size=9, bold=True, color=theme.TEXT_MUTED),
        fill=PatternFill("solid", fgColor=theme.BG_SUBTLE),
        align=Alignment(vertical="center", wrap_text=True),
    )
    add("x-cell")
    add("x-cell-muted", font=Font(name=FONT, size=10, color=theme.TEXT_MUTED))
    add("x-num", align=right)
    add("x-date", number_format="DD.MM.YYYY", align=right)
    add("x-pct", number_format="0%", align=right)
    # Отклонение со знаком: «+4» читается как отставание, «4» — как просто
    # число дней.
    add("x-dev", number_format="+0;-0;0", align=right)
    add(
        "x-group",
        font=Font(name=FONT, size=9, bold=True, color=theme.ACCENT),
        fill=PatternFill("solid", fgColor=theme.ACCENT_SOFT),
    )


def _widths(ws: Worksheet, widths: list[float]) -> None:
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width


def _fill(color: str) -> PatternFill:
    return PatternFill("solid", fgColor=color)


def _sheet(wb: Workbook, title: str) -> Worksheet:
    # Excel режет имя листа на 31 знаке и падает на нескольких запрещённых
    # знаках. Обрезаем сами: молчаливое усечение библиотекой дало бы два листа
    # с одинаковым именем на длинных переводах.
    safe = "".join(ch for ch in title if ch not in r"[]:*?/\\")[:31]
    ws = wb.create_sheet(safe)
    ws.sheet_view.showGridLines = False
    return ws


# --- Обзор --------------------------------------------------------------------


def _overview(wb: Workbook, doc: ExportDocument) -> None:
    t = doc.labels
    ws = _sheet(wb, t("section", "overview"))
    # Колонка A — узкий жёлоб под образец цвета категории. Плитки начинаются с
    # B: в узкой колонке Excel показывает вместо числа «#».
    _widths(ws, [3, 20, 18, 18, 18, 18, 18, 26])

    ws["B2"] = doc.project_name
    ws["B2"].style = "x-h1"
    ws.row_dimensions[2].height = 30
    ws["B3"] = f"{doc.org_name}  ·  {doc.plan_badge}"
    ws["B3"].style = "x-muted"

    period = f"{t('doc', 'period')}: {doc.start:%d.%m.%Y} — {doc.end:%d.%m.%Y}"
    if doc.deadline:
        period += f"   ·   {t('doc', 'deadline')}: {doc.deadline:%d.%m.%Y}"
    period += f"   ·   {t('doc', 'generated')}: {doc.generated_at:%d.%m.%Y}"
    ws["B4"] = period
    ws["B4"].style = "x-muted"
    if doc.client_copy:
        ws["B5"] = t("doc", "client_copy")
        ws["B5"].style = "x-muted"

    kpi = doc.kpi
    tiles = [
        (t("kpi", "total"), kpi.total, theme.TAG_GRAY, theme.TEXT, None),
        (t("kpi", "done"), kpi.done, theme.OK_SOFT, theme.OK, None),
        (t("kpi", "in_progress"), kpi.in_progress, theme.ACCENT_SOFT, theme.ACCENT, None),
        (t("kpi", "blocked"), kpi.blocked, theme.DANGER_SOFT, theme.DANGER_STRONG, None),
        (t("kpi", "late"), kpi.late, theme.WARN_SOFT, theme.WARN_TEXT, None),
        (t("kpi", "progress"), kpi.progress_pct / 100, theme.TAG_GRAY, theme.TEXT, "0%"),
    ]
    ws.row_dimensions[7].height = 30
    ws.row_dimensions[8].height = 18
    for i, (name, value, bg, fg, fmt) in enumerate(tiles):
        column = get_column_letter(2 + i)
        cell = ws[f"{column}7"]
        cell.value = value
        cell.font = Font(name=FONT, size=18, bold=True, color=fg)
        cell.fill = _fill(bg)
        cell.alignment = Alignment(vertical="center", horizontal="center")
        if fmt:
            cell.number_format = fmt
        caption = ws[f"{column}8"]
        caption.value = name
        caption.font = Font(name=FONT, size=9, color=theme.TEXT_MUTED)
        caption.fill = _fill(bg)
        caption.alignment = Alignment(vertical="center", horizontal="center")

    row = 11
    ws.cell(row=row, column=2, value=t("col", "category")).style = "x-h2"
    row += 1
    for category in doc.categories:
        ws.cell(row=row, column=1).fill = _fill(category.color.lstrip("#").upper())
        ws.cell(row=row, column=2, value=category.name).style = "x-cell"
        ws.cell(
            row=row, column=3, value=t("count", "tasks", n=category.task_count)
        ).style = "x-cell-muted"
        row += 1


# --- Задачи -------------------------------------------------------------------


def _tasks(wb: Workbook, doc: ExportDocument) -> None:
    t = doc.labels
    ws = _sheet(wb, t("section", "tasks"))

    show_notes = any(task.note for task in doc.tasks)
    show_people = any(task.assignees for task in doc.tasks)
    show_baseline = not doc.client_copy

    columns: list[tuple[str, float, str]] = [
        (t("col", "n"), 5, "x-num"),
        (t("col", "task"), 38, "x-cell"),
        (t("col", "status"), 15, "x-cell"),
        (t("col", "criticality"), 12, "x-cell"),
        (t("col", "progress"), 11, "x-pct"),
        (t("col", "start"), 12, "x-date"),
        (t("col", "end"), 12, "x-date"),
        (t("col", "days"), 7, "x-num"),
        (t("col", "milestone"), 8, "x-cell"),
        (t("col", "critical"), 8, "x-cell"),
    ]
    if show_people:
        columns.append((t("col", "assignees"), 26, "x-cell"))
    if show_baseline:
        columns += [
            (t("col", "baseline_start"), 14, "x-date"),
            (t("col", "baseline_end"), 14, "x-date"),
            (t("col", "deviation"), 10, "x-dev"),
        ]
    if show_notes:
        columns.append((t("col", "note"), 42, "x-cell-muted"))

    _widths(ws, [width for _, width, _ in columns])
    ws.row_dimensions[1].height = 26
    for i, (name, _, _) in enumerate(columns, start=1):
        ws.cell(row=1, column=i, value=name).style = "x-head"

    def values_of(task: DocTask) -> list[object]:
        out: list[object] = [
            task.number,
            ("◆ " if task.milestone else "") + task.name,
            task.status_label,
            task.criticality_label,
            task.progress_pct / 100,
            task.start,
            task.end,
            task.duration_days,
            "◆" if task.milestone else "",
            "•" if task.critical else "",
        ]
        if show_people:
            out.append(", ".join(task.assignees))
        if show_baseline:
            out += [task.baseline_start, task.baseline_end, task.deviation_days]
        if show_notes:
            out.append(task.note)
        return out

    row = 2
    for category in doc.categories:
        for i in range(1, len(columns) + 1):
            ws.cell(
                row=row, column=i, value=category.name if i == 2 else None
            ).style = "x-group"
        ws.row_dimensions[row].height = ROW_H
        row += 1
        for task in doc.tasks_of(category.id):
            for i, (value, (_, _, style)) in enumerate(
                zip(values_of(task), columns), start=1
            ):
                cell = ws.cell(row=row, column=i, value=value)
                cell.style = style
                if i in (9, 10):
                    cell.alignment = Alignment(vertical="center", horizontal="center")

            name_cell = ws.cell(row=row, column=2)
            name_cell.font = Font(name=FONT, size=10, bold=True, color=theme.TEXT)

            status_cell = ws.cell(row=row, column=3)
            bg, fg = theme.STATUS_CHIP[task.status]
            status_cell.fill = _fill(bg)
            status_cell.font = Font(name=FONT, size=9, bold=True, color=fg)
            status_cell.alignment = Alignment(vertical="center", horizontal="center")

            if task.criticality == "critical":
                ws.cell(row=row, column=4).font = Font(
                    name=FONT, size=9, bold=True, color=theme.DANGER_STRONG
                )
            # Строки задач сворачиваются родным «плюсом» Excel по категориям.
            ws.row_dimensions[row].outlineLevel = 1
            ws.row_dimensions[row].height = ROW_H
            row += 1

    last = row - 1
    if last < 2:
        return

    ws.freeze_panes = "C2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{last}"
    # Итог сворачивания — строка категории сверху, а не снизу: она заголовок,
    # а не сумма.
    ws.sheet_properties.outlinePr.summaryBelow = False

    ws.conditional_formatting.add(
        f"E2:E{last}",
        DataBarRule(
            start_type="num",
            start_value=0,
            end_type="num",
            end_value=1,
            color=theme.ACCENT,
            showValue=True,
        ),
    )
    # Просрочка — правилом по колонке конца, а не отдельным столбцом с
    # восклицательным знаком: правило переживает сортировку и фильтр, столбец
    # с готовым ответом — нет.
    if doc.dated:
        today = doc.today
        ws.conditional_formatting.add(
            f"G2:G{last}",
            CellIsRule(
                operator="lessThan",
                formula=[f"DATE({today.year},{today.month},{today.day})"],
                fill=_fill(theme.WARN_SOFT),
                font=Font(name=FONT, size=10, color=theme.WARN_TEXT),
            ),
        )


# --- Диаграмма Ганта ----------------------------------------------------------


def _gantt(wb: Workbook, doc: ExportDocument) -> None:
    t = doc.labels
    ws = _sheet(wb, t("section", "gantt"))

    window = doc.layout.window
    zoom = budget.default_zoom_for_xlsx(window.days)
    budget.require_within_xlsx_budget(window, zoom)
    step = budget.DAYS_PER_UNIT[zoom]
    columns = budget.columns_for(window.days, zoom)

    LEFT = 3  # задача | начало | конец
    _widths(ws, [34, 11, 11] + [DAY_COL_W] * columns)

    def column_of(day: date) -> int:
        return LEFT + 1 + (day - window.start).days // step

    def day_of(index: int) -> date:
        return window.start + timedelta(days=index * step)

    ws.cell(row=1, column=1, value=doc.project_name).style = "x-h2"
    ws.cell(row=2, column=1, value=t("col", "task")).style = "x-head"
    ws.cell(row=2, column=2, value=t("col", "start")).style = "x-head"
    ws.cell(row=2, column=3, value=t("col", "end")).style = "x-head"

    _scale_header(ws, doc, window, zoom, columns, LEFT, day_of)

    weekend = _fill(theme.NONWORKING)
    today_fill = _fill(theme.TODAY_CELL)
    ghost = _fill(theme.BASELINE_GHOST)

    row = 3
    for category in doc.categories:
        ws.cell(row=row, column=1, value=category.name.upper()).style = "x-group"
        for i in range(2, LEFT + 1 + columns):
            ws.cell(row=row, column=i).fill = _fill(theme.ACCENT_SOFT)
        ws.row_dimensions[row].height = ROW_H
        row += 1

        for task in doc.tasks_of(category.id):
            ws.cell(
                row=row,
                column=1,
                value=("◆ " if task.milestone else "   ") + task.name,
            ).style = "x-cell"
            ws.cell(row=row, column=2, value=task.start).style = "x-date"
            ws.cell(row=row, column=3, value=task.end).style = "x-date"
            ws.row_dimensions[row].height = ROW_H

            # Подложка: нерабочие дни и сегодня. Только в дневном масштабе —
            # колонка недели или месяца не бывает выходной целиком.
            if zoom is Zoom.DAY:
                for i in range(columns):
                    day = day_of(i)
                    cell = ws.cell(row=row, column=LEFT + 1 + i)
                    if doc.dated and day == doc.today:
                        cell.fill = today_fill
                    elif day.weekday() >= 5:
                        cell.fill = weekend

            if task.baseline_start and task.baseline_end and not task.milestone:
                for column in range(
                    column_of(max(task.baseline_start, window.start)),
                    column_of(min(task.baseline_end, window.end)) + 1,
                ):
                    ws.cell(row=row, column=column).fill = ghost

            _bar(ws, row, task, window, column_of)
            row += 1

    ws.freeze_panes = ws.cell(row=3, column=LEFT + 1)


def _scale_header(ws, doc, window, zoom, columns, left: int, day_of) -> None:
    """Две строки шапки: месяцы объединённой строкой, под ними — числа."""
    t = doc.labels
    run_start = 0
    for i in range(columns + 1):
        current = day_of(i) if i < columns else None
        boundary = current is None or current.month != day_of(run_start).month
        if not boundary:
            continue
        first, last = left + 1 + run_start, left + i
        if last > first:
            ws.merge_cells(start_row=1, start_column=first, end_row=1, end_column=last)
        month = day_of(run_start)
        cell = ws.cell(row=1, column=first, value=f"{t.month(month.month)} {month.year}")
        cell.font = Font(name=FONT, size=9, bold=True, color=theme.TEXT)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = _fill(theme.BG_SUBTLE)
        run_start = i

    for i in range(columns):
        day = day_of(i)
        # В дневном масштабе числа через одно: подряд они слипаются на ширине
        # в три знака.
        text = str(day.day) if zoom is not Zoom.DAY or day.day % 2 else None
        cell = ws.cell(row=2, column=left + 1 + i, value=text)
        cell.font = Font(name=FONT, size=7, color=theme.TEXT_FAINT)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = _fill(theme.BG_SUBTLE)


def _bar(ws, row: int, task: DocTask, window, column_of) -> None:
    """Полоска задачи ячейками. Веха — ромб в своей колонке, а не отрезок."""
    if task.end < window.start or task.start > window.end:
        return

    if task.milestone:
        cell = ws.cell(row=row, column=column_of(task.start), value="◆")
        cell.font = Font(name=FONT, size=10, bold=True, color=theme.TEXT)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        return

    first = column_of(max(task.start, window.start))
    last = column_of(min(task.end, window.end))
    fill = _fill(theme.STATUS_BAR[task.status][0])

    # Запланированная полоска по правилу ленты не залита — а бледная заливка
    # на белом почти не видна и на призраке базового плана не видна вовсе.
    # Поэтому в книге она ещё и обводится.
    edge = Side(style="thin", color=theme.LINE_PLANNED) if task.status == "planned" else None
    warn = Side(style="medium", color=theme.WARN) if task.late else None
    outline = warn or edge

    for column in range(first, last + 1):
        cell = ws.cell(row=row, column=column)
        cell.fill = fill
        if outline is not None:
            cell.border = Border(
                top=outline,
                bottom=outline,
                left=outline if column == first else None,
                right=outline if column == last else None,
            )


# --- прочие листы -------------------------------------------------------------


def _links(wb: Workbook, doc: ExportDocument) -> None:
    t = doc.labels
    ws = _sheet(wb, t("section", "links"))
    _widths(ws, [38, 38, 14, 14])
    head = [t("col", "from"), t("col", "to"), t("col", "end"), t("col", "start")]
    for i, name in enumerate(head, start=1):
        ws.cell(row=1, column=i, value=name).style = "x-head"

    for row, link in enumerate(doc.links, start=2):
        ws.cell(row=row, column=1, value=link.from_name).style = "x-cell"
        ws.cell(row=row, column=2, value=link.to_name).style = "x-cell"
        ws.cell(row=row, column=3, value=link.from_end).style = "x-date"
        ws.cell(row=row, column=4, value=link.to_start).style = "x-date"
        if link.broken:
            for i in range(1, 5):
                ws.cell(row=row, column=i).fill = _fill(theme.WARN_SOFT)
    ws.freeze_panes = "A2"


def _proposal(wb: Workbook, doc: ExportDocument) -> None:
    t = doc.labels
    proposal = doc.proposal
    ws = _sheet(wb, t("section", "proposal"))
    _widths(ws, [40, 26, 10, 10, 16])
    money = f'# ##0 "{proposal.currency}"'
    head = [
        t("col", "task"),
        t("col", "role"),
        t("col", "effort"),
        t("col", "rate"),
        t("col", "sum"),
    ]
    for i, name in enumerate(head, start=1):
        ws.cell(row=1, column=i, value=name).style = "x-head"

    row = 2
    amounts: list[int] = []
    for group in proposal.groups:
        for i in range(1, 6):
            ws.cell(row=row, column=i, value=group.name if i == 1 else None).style = "x-group"
        row += 1
        for line in group.lines:
            ws.cell(row=row, column=1, value=line.name).style = "x-cell"
            ws.cell(row=row, column=2, value=line.role).style = "x-cell-muted"
            ws.cell(row=row, column=3, value=float(line.effort)).style = "x-num"
            ws.cell(row=row, column=4, value=float(line.rate)).style = "x-num"
            # Цена — формулой, а не числом: получатель правит ставку и видит
            # новый итог, не заказывая выгрузку заново. Это единственное место
            # в книге, где формулы заслуживают своего усложнения.
            cell = ws.cell(row=row, column=5, value=f"=C{row}*D{row}")
            cell.style = "x-num"
            cell.number_format = money
            amounts.append(row)
            row += 1

    if not amounts:
        return

    total = "+".join(f"E{i}" for i in amounts)
    row += 1
    lines = [
        (t("total", "subtotal"), f"={total}", False),
        (t("total", "tax", p=_plain(proposal.tax_rate_pct)), f"=E{row}*{proposal.tax_rate_pct}/100", False),
        (t("total", "total"), f"=E{row}+E{row + 1}", True),
    ]
    for name, formula, bold in lines:
        label = ws.cell(row=row, column=4, value=name)
        label.alignment = Alignment(horizontal="right", vertical="center")
        label.font = Font(
            name=FONT, size=10, bold=bold, color=theme.TEXT if bold else theme.TEXT_MUTED
        )
        cell = ws.cell(row=row, column=5, value=formula)
        cell.number_format = money
        cell.alignment = Alignment(horizontal="right", vertical="center")
        cell.font = Font(name=FONT, size=11 if bold else 10, bold=bold, color=theme.TEXT)
        row += 1
    ws.freeze_panes = "A2"


def _scorecard(wb: Workbook, doc: ExportDocument) -> None:
    t = doc.labels
    card = doc.scorecard
    ws = _sheet(wb, t("section", "scorecard"))
    _widths(ws, [30, 20, 9, 9] + [9] * len(card.weeks))

    head = [t("col", "metric"), t("col", "owner"), t("col", "target"), ""]
    head += [f"{week:%d.%m}" for week in card.weeks]
    for i, name in enumerate(head, start=1):
        ws.cell(row=1, column=i, value=name).style = "x-head"

    for row, metric in enumerate(card.metrics, start=2):
        ws.cell(row=row, column=1, value=metric.label).style = "x-cell"
        ws.cell(row=row, column=2, value=metric.owner or "—").style = "x-cell-muted"
        ws.cell(row=row, column=3, value=metric.target).style = "x-num"
        ws.cell(
            row=row, column=4, value="≤" if metric.direction == "lte" else "≥"
        ).style = "x-cell-muted"
        for i, (value, status) in enumerate(zip(metric.values, metric.statuses)):
            cell = ws.cell(row=row, column=5 + i, value=value if value is not None else "—")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            bg, fg = theme.METRIC_CELL[status]
            cell.fill = _fill(bg)
            cell.font = Font(name=FONT, size=9, bold=status != "no_data", color=fg)
    ws.freeze_panes = "E2"


def _comments(wb: Workbook, doc: ExportDocument) -> None:
    t = doc.labels
    ws = _sheet(wb, t("section", "comments"))
    _widths(ws, [26, 18, 12, 12, 70])
    head = [
        t("col", "task"),
        t("col", "author"),
        t("col", "date"),
        t("col", "visibility"),
        t("col", "text"),
    ]
    for i, name in enumerate(head, start=1):
        ws.cell(row=1, column=i, value=name).style = "x-head"

    for row, comment in enumerate(doc.comments, start=2):
        ws.cell(row=row, column=1, value=comment.task_name or "—").style = "x-cell"
        ws.cell(row=row, column=2, value=comment.author).style = "x-cell"
        ws.cell(row=row, column=3, value=comment.created_at).style = "x-date"
        cell = ws.cell(
            row=row,
            column=4,
            value=t("comment", "internal" if comment.internal else "public"),
        )
        cell.style = "x-cell-muted"
        if comment.internal:
            cell.fill = _fill(theme.WARN_SOFT)
            cell.font = Font(name=FONT, size=9, bold=True, color=theme.WARN_TEXT)
        ws.cell(row=row, column=5, value=comment.body).style = "x-cell-muted"
    ws.freeze_panes = "A2"


def _history(wb: Workbook, doc: ExportDocument) -> None:
    t = doc.labels
    ws = _sheet(wb, t("section", "history"))
    _widths(ws, [14, 22, 34, 40])
    head = [t("col", "date"), t("col", "author"), t("col", "event"), t("col", "task")]
    for i, name in enumerate(head, start=1):
        ws.cell(row=1, column=i, value=name).style = "x-head"

    for row, event in enumerate(doc.history, start=2):
        ws.cell(row=row, column=1, value=event.at).style = "x-date"
        ws.cell(row=row, column=2, value=event.actor or "—").style = "x-cell"
        ws.cell(row=row, column=3, value=event.event).style = "x-cell"
        ws.cell(row=row, column=4, value=event.subject).style = "x-cell-muted"
    ws.freeze_panes = "A2"


def _plain(value) -> str:
    """Процент без хвоста нулей: «18», а не «18.00»."""
    text = f"{value:f}".rstrip("0").rstrip(".")
    return text or "0"


# --- сборка -------------------------------------------------------------------

#: Какой лист рисует какая функция. Порядок листов — порядок словаря: обзор
#: первым, журнал последним, как их и читают.
_SHEETS = [
    (ExportSection.OVERVIEW, _overview),
    (ExportSection.TASKS, _tasks),
    (ExportSection.GANTT, _gantt),
    (ExportSection.LINKS, _links),
    (ExportSection.PROPOSAL, _proposal),
    (ExportSection.SCORECARD, _scorecard),
    (ExportSection.COMMENTS, _comments),
    (ExportSection.HISTORY, _history),
]

#: Разделы, у которых нет данных, лист не заводят вовсе: страница из одних
#: заголовков читается как поломка выгрузки, а не как «здесь пусто».
def _has_content(doc: ExportDocument, section: ExportSection) -> bool:
    if section is ExportSection.PROPOSAL:
        return doc.proposal is not None
    if section is ExportSection.SCORECARD:
        return doc.scorecard is not None
    if section is ExportSection.COMMENTS:
        return bool(doc.comments)
    if section is ExportSection.HISTORY:
        return bool(doc.history)
    if section is ExportSection.LINKS:
        return bool(doc.links)
    return True


def render(doc: ExportDocument) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)
    _styles(wb)

    for section, draw in _SHEETS:
        if doc.has(section) and _has_content(doc, section):
            draw(wb, doc)

    # Книга без единого листа Excel не открывает вовсе — а пустой выбор
    # отсекается маршрутом раньше, так что сюда мы попадаем только если все
    # выбранные разделы оказались пустыми.
    if not wb.worksheets:
        _overview(wb, doc)

    # Печать: альбомная и «вписать по ширине» на каждом листе. Книгу, которую
    # получатель первым делом печатает, стыдно отдавать разорванной по колонкам.
    for ws in wb.worksheets:
        ws.page_setup.orientation = "landscape"
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.print_options.horizontalCentered = True
        ws.page_margins.left = ws.page_margins.right = 0.3
        ws.page_margins.top = ws.page_margins.bottom = 0.4

    wb.properties.title = doc.project_name
    wb.properties.creator = "Planora"

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
