"""Выгрузка проекта в Excel и PDF.

Проверяется не «двести не пятьсот», а три обещания, ради которых работа
затевалась: файл открывается своим приложением и содержит настоящие даты и
числа; клиенту по ссылке не уезжает то, что ему не обещано; и число страниц
ленты не растёт молча — правило масштаба отказывает раньше, чем соберёт
неподъёмный документ.
"""

import io
import json
from datetime import date, timedelta
from pathlib import Path

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook

from app.db import get_db
from app.export import budget, theme
from app.export.budget import Orientation, Period, Zoom
from app.export.document import align_weeks, metric_row
from app.export.labels import available_locales, dictionary
from app.main import app

XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

ALL = "include=overview&include=tasks&include=gantt&include=links&include=proposal&include=scorecard&include=comments&include=history"


@pytest.fixture
def client(db):
    def _override_get_db():
        yield db

    app.dependency_overrides[get_db] = _override_get_db
    try:
        yield TestClient(app)
    finally:
        app.dependency_overrides.pop(get_db, None)


@pytest.fixture(autouse=True)
def fresh_rate_limit(monkeypatch):
    """Счётчик выгрузок живёт в памяти процесса: без сброса один тест доедал
    бы окно следующего (тот же приём, что у гостевых комментариев)."""
    import app.api.export_routes as export_routes

    monkeypatch.setattr(export_routes, "_limiter", None)


@pytest.fixture
def authed(client):
    client.post(
        "/api/auth/register",
        json={
            "name": "Alex",
            "email": "alex@example.com",
            "password": "s3cret-pass",
            "company_name": "Acme",
        },
    )
    return client


def _mutate(authed, project_id, op):
    response = authed.post(f"/api/projects/{project_id}/mutations", json={"op": op})
    assert response.status_code in (200, 201), response.text
    return response.json()


def _make_project(authed, *, tasks=3, start=date(2026, 3, 2), duration=5, name="Redesign"):
    project_id = authed.post("/api/projects", json={"name": name}).json()["id"]
    # Дата старта назначается до задач: в календарном режиме задача создаётся
    # с настоящей датой, а в относительном — координатой на оси «День N», и
    # привязка задним числом переразложила бы уже созданные задачи по
    # рабочим дням.
    anchored = authed.post(
        f"/api/projects/{project_id}/schedule", json={"start_date": start.isoformat()}
    )
    assert anchored.status_code in (200, 201), anchored.text
    category_id = _mutate(
        authed, project_id, {"type": "create_category", "name": "Дизайн", "color": "#3b82f6"}
    )["op"]["category_id"]

    task_ids = []
    for i in range(tasks):
        task_ids.append(
            _mutate(
                authed,
                project_id,
                {
                    "type": "create_task",
                    "category_id": category_id,
                    "name": f"Задача {i + 1}",
                    "start_date": (start + timedelta(days=i * duration)).isoformat(),
                    "duration_days": duration,
                },
            )["op"]["task_id"]
        )
    return project_id, category_id, task_ids


# --- форма ответа -------------------------------------------------------------


def test_both_formats_come_back_as_files_with_a_name(authed):
    project_id, _, _ = _make_project(authed)

    for fmt, mime, signature in (
        ("xlsx", XLSX_MIME, b"PK"),
        ("pdf", "application/pdf", b"%PDF-"),
    ):
        response = authed.get(f"/api/projects/{project_id}/export.{fmt}?{ALL}")
        assert response.status_code == 200, response.text
        assert response.headers["content-type"].startswith(mime)
        assert response.content.startswith(signature)

        # Имя файла — дважды: ASCII-заглушка и RFC 5987. Без второго кириллица
        # в имени проекта не переживает заголовок.
        disposition = response.headers["content-disposition"]
        assert disposition.startswith("attachment;")
        assert f".{fmt}" in disposition
        assert "filename*=UTF-8''" in disposition
        # Файл собран под права спрашивающего — общему кэшу его отдавать нельзя.
        assert "no-store" in response.headers["cache-control"]


def test_a_cyrillic_project_name_survives_the_header(authed):
    project_id, _, _ = _make_project(authed, name="Переезд офиса")
    response = authed.get(f"/api/projects/{project_id}/export.pdf?{ALL}")
    assert response.status_code == 200

    disposition = response.headers["content-disposition"]
    # В ASCII-части кириллицы быть не может, зато она обязана быть в
    # процентной кодировке — иначе браузер сохранит файл под «____».
    assert "%D0%9F%D0%B5%D1%80%D0%B5%D0%B5%D0%B7%D0%B4" in disposition
    assert disposition.split(";")[1].strip().isascii()


# --- книга Excel --------------------------------------------------------------


def test_the_workbook_carries_real_dates_and_numbers_not_strings(authed):
    """Даты строками превратили бы книгу в картинку таблицы: Excel не
    сортирует и не фильтрует по ним."""
    project_id, _, task_ids = _make_project(authed, start=date(2026, 3, 2))
    _mutate(
        authed, project_id, {"type": "set_progress", "task_id": task_ids[0], "progress_pct": 40}
    )
    body = authed.get(f"/api/projects/{project_id}/export.xlsx?{ALL}&locale=ru").content

    wb = load_workbook(io.BytesIO(body))
    sheet = wb["Задачи"]
    header = [cell.value for cell in sheet[1]]
    start_column = header.index("Начало") + 1
    progress_column = header.index("Прогресс") + 1

    # Первая строка данных — заголовок категории; задача идёт следом.
    row = 3
    # Excel хранит дату числом со временем; openpyxl отдаёт её datetime — это
    # и есть настоящая дата, а не строка, по которой нельзя сортировать.
    assert sheet.cell(row=row, column=start_column).value.date() == date(2026, 3, 2)
    assert sheet.cell(row=row, column=start_column).number_format == "DD.MM.YYYY"

    # Прогресс — доля с процентным форматом, а не текст «40%»: иначе по
    # колонке не построить ни фильтр, ни среднее.
    assert sheet.cell(row=row, column=progress_column).value == pytest.approx(0.4)
    assert sheet.cell(row=row, column=progress_column).number_format == "0%"


def test_the_workbook_has_exactly_the_requested_sheets(authed):
    project_id, _, _ = _make_project(authed)

    body = authed.get(
        f"/api/projects/{project_id}/export.xlsx?include=tasks&include=gantt&locale=ru"
    ).content
    assert [ws.title for ws in load_workbook(io.BytesIO(body)).worksheets] == [
        "Задачи",
        "Диаграмма Ганта",
    ]


def test_an_empty_section_does_not_produce_an_empty_sheet(authed):
    """Лист из одних заголовков читается как поломка выгрузки, а не как
    «здесь пусто»."""
    project_id, _, _ = _make_project(authed)
    body = authed.get(
        f"/api/projects/{project_id}/export.xlsx?include=tasks&include=comments&locale=ru"
    ).content
    assert [ws.title for ws in load_workbook(io.BytesIO(body)).worksheets] == ["Задачи"]


def test_the_status_cell_is_painted_by_the_same_palette_as_the_chart(authed):
    project_id, _, task_ids = _make_project(authed)
    _mutate(authed, project_id, {"type": "set_status", "task_id": task_ids[0], "status": "done"})

    body = authed.get(f"/api/projects/{project_id}/export.xlsx?include=tasks&locale=ru").content
    sheet = load_workbook(io.BytesIO(body))["Задачи"]
    header = [cell.value for cell in sheet[1]]
    status_column = header.index("Статус") + 1

    cell = sheet.cell(row=3, column=status_column)
    assert cell.value == "Готово"
    assert cell.fill.fgColor.rgb.endswith("E9F8F0")  # theme.OK_SOFT


def test_the_proposal_totals_are_formulas_so_a_rate_can_be_edited(authed):
    project_id, _, _ = _make_project(authed)
    category_id = authed.post(
        f"/api/projects/{project_id}/proposal/categories", json={"name": "Работы"}
    ).json()["id"]
    authed.post(
        f"/api/projects/{project_id}/proposal/categories/{category_id}/tasks",
        json={"name": "Вёрстка", "role": "Дизайнер", "effort": 4, "rate": 100},
    )

    body = authed.get(f"/api/projects/{project_id}/export.xlsx?include=proposal&locale=ru").content
    sheet = load_workbook(io.BytesIO(body))["Смета"]
    formulas = [
        cell.value
        for row in sheet.iter_rows()
        for cell in row
        if isinstance(cell.value, str) and cell.value.startswith("=")
    ]
    assert any("*" in formula for formula in formulas), "цена строки — формула"
    assert any("+" in formula for formula in formulas), "итог — формула"


# --- документ PDF -------------------------------------------------------------


def _pdf_text(body: bytes) -> str:
    from pypdf import PdfReader

    reader = PdfReader(io.BytesIO(body))
    return "\n".join(page.extract_text() or "" for page in reader.pages)


def _pdf_pages(body: bytes) -> int:
    from pypdf import PdfReader

    return len(PdfReader(io.BytesIO(body)).pages)


def test_the_pdf_names_the_project_and_its_tasks(authed):
    project_id, _, _ = _make_project(authed, name="Переезд офиса")
    body = authed.get(f"/api/projects/{project_id}/export.pdf?{ALL}&locale=ru").content

    text = _pdf_text(body)
    assert "Переезд офиса" in text
    assert "Задача 1" in text


@pytest.mark.parametrize("locale,word", [("az", "Tapşırıqlar"), ("en", "Tasks"), ("ru", "Задачи")])
def test_the_pdf_speaks_the_asked_language(authed, locale, word):
    """Ради этого и встраивается Inter: во встроенных шрифтах ReportLab нет ни
    кириллицы, ни `ə`."""
    project_id, _, _ = _make_project(authed)
    body = authed.get(
        f"/api/projects/{project_id}/export.pdf?include=tasks&locale={locale}"
    ).content
    assert word in _pdf_text(body)


def test_a_wide_project_is_split_by_time_not_cropped(authed):
    """Проект шире страницы становится несколькими страницами ленты — и на
    каждой повторяется колонка названий."""
    project_id, _, _ = _make_project(authed, tasks=12, duration=20)
    body = authed.get(
        f"/api/projects/{project_id}/export.pdf?include=gantt&zoom=day&locale=ru"
    ).content

    assert _pdf_pages(body) > 1
    # Подпись первой задачи стоит на каждой странице ленты, а не только там,
    # где лежит её полоска.
    from pypdf import PdfReader

    pages = [page.extract_text() or "" for page in PdfReader(io.BytesIO(body)).pages]
    assert all("Задача 1" in page for page in pages)


# --- клиентский экземпляр ------------------------------------------------------


@pytest.fixture
def shared(authed):
    """Проект с внутренней заметкой, внутренней репликой и публичной ссылкой."""
    project_id, category_id, task_ids = _make_project(authed, name="Публичный")
    _mutate(
        authed,
        project_id,
        {
            "type": "set_task_fields",
            "task_id": task_ids[0],
            "name": "Задача 1",
            "description": "",
            "internal_note": "СЕКРЕТ-ЗАМЕТКА",
        },
    )
    authed.post(
        f"/api/projects/{project_id}/comments",
        json={"body": "СЕКРЕТ-РЕПЛИКА", "internal": True},
    )
    authed.post(f"/api/projects/{project_id}/comments", json={"body": "Общая реплика"})
    link = authed.post(f"/api/projects/{project_id}/share").json()
    return project_id, link


def _link_parts(link: dict) -> tuple[str, str, str]:
    """Слаги и строка запроса из публичной ссылки.

    Имя параметра токена берётся из самой ссылки, а не пишется здесь руками:
    его знает `app.sharing.TOKEN_PARAM`, и тест, повторивший это знание,
    разошёлся бы с ним молча.
    """
    from urllib.parse import urlparse

    parsed = urlparse(link["url"])
    org_slug, project_slug = parsed.path.split("/p/")[-1].split("/")[:2]
    return org_slug, project_slug, parsed.query


def test_the_guest_copy_hides_notes_people_and_the_baseline(authed, shared, client):
    project_id, link = shared
    authed.post(f"/api/projects/{project_id}/plan/approvals")

    org_slug, project_slug, query = _link_parts(link)

    response = client.get(
        f"/api/public/{org_slug}/{project_slug}/export.xlsx?{ALL}&locale=ru&{query}"
    )
    assert response.status_code == 200, response.text
    wb = load_workbook(io.BytesIO(response.content))

    # Выгрузка не показывает больше, чем показывает страница, с которой её
    # позвали: ни сметы со ставками, ни скоркарда, ни журнала правок публичная
    # страница гостю не отдаёт.
    assert "История правок" not in wb.sheetnames
    assert "Смета" not in wb.sheetnames
    assert "Скоркард" not in wb.sheetnames

    header = [cell.value for cell in wb["Задачи"][1]]
    assert "Заметка" not in header
    assert "Исполнители" not in header
    assert "База: начало" not in header

    dump = json.dumps(
        [[cell.value for cell in row] for ws in wb.worksheets for row in ws.iter_rows()],
        ensure_ascii=False,
        default=str,
    )
    assert "СЕКРЕТ-ЗАМЕТКА" not in dump
    assert "СЕКРЕТ-РЕПЛИКА" not in dump


def test_the_member_copy_still_carries_what_the_guest_may_not_see(authed, shared):
    """Обратная сторона предыдущего теста: без неё он проходил бы и на
    выгрузке, потерявшей заметки для всех."""
    project_id, _ = shared
    category_id = authed.post(
        f"/api/projects/{project_id}/proposal/categories", json={"name": "Работы"}
    ).json()["id"]
    authed.post(
        f"/api/projects/{project_id}/proposal/categories/{category_id}/tasks",
        json={"name": "Вёрстка", "role": "Дизайнер", "effort": 4, "rate": 100},
    )

    body = authed.get(f"/api/projects/{project_id}/export.xlsx?{ALL}&locale=ru").content
    wb = load_workbook(io.BytesIO(body))

    assert "История правок" in wb.sheetnames
    assert "Смета" in wb.sheetnames
    dump = json.dumps(
        [[cell.value for cell in row] for ws in wb.worksheets for row in ws.iter_rows()],
        ensure_ascii=False,
        default=str,
    )
    assert "СЕКРЕТ-ЗАМЕТКА" in dump
    assert "СЕКРЕТ-РЕПЛИКА" in dump


# --- отказы --------------------------------------------------------------------


def test_an_unknown_section_is_refused_by_the_schema(authed):
    project_id, _, _ = _make_project(authed)
    response = authed.get(f"/api/projects/{project_id}/export.pdf?include=payroll")
    assert response.status_code == 422


def test_an_empty_selection_is_refused_rather_than_silently_meaning_everything(authed):
    project_id, _, _ = _make_project(authed)
    response = authed.get(f"/api/projects/{project_id}/export.pdf")
    assert response.status_code == 422
    assert response.json()["detail"] == "export_empty_selection"


def test_a_project_beyond_the_task_ceiling_is_refused_before_the_work_starts(
    authed, monkeypatch
):
    from app.api import export_routes

    monkeypatch.setattr(export_routes, "MAX_TASKS", 1)
    project_id, _, _ = _make_project(authed, tasks=3)
    response = authed.get(f"/api/projects/{project_id}/export.pdf?include=tasks")
    assert response.status_code == 422
    assert response.json()["detail"] == "export_too_large"


# --- правило масштаба ----------------------------------------------------------


@pytest.mark.parametrize(
    "months,expected",
    [(3, Zoom.DAY), (6, Zoom.WEEK), (12, Zoom.WEEK), (24, Zoom.MONTH)],
)
def test_the_default_zoom_is_the_most_detailed_that_still_fits(months, expected):
    days = months * 30
    assert budget.default_zoom(days, Orientation.LANDSCAPE) is expected
    assert budget.page_count(days, expected, Orientation.LANDSCAPE) <= budget.COMFORTABLE_PAGES


def test_the_server_picks_a_zoom_when_the_dialog_did_not(authed):
    """Маршрут зовут и мимо окна — закладкой, скриптом, публичной ссылкой."""
    project_id, _, _ = _make_project(authed, tasks=4, duration=30)
    response = authed.get(f"/api/projects/{project_id}/export.pdf?include=gantt")
    assert response.status_code == 200
    assert _pdf_pages(response.content) <= budget.COMFORTABLE_PAGES


def test_a_scale_beyond_the_ceiling_is_refused_instead_of_forty_pages(authed):
    project_id, _, _ = _make_project(authed, tasks=20, duration=60)
    response = authed.get(
        f"/api/projects/{project_id}/export.pdf?include=gantt&zoom=day"
    )
    assert response.status_code == 422
    assert response.json()["detail"] == "export_scale_too_wide"


def test_narrowing_the_period_brings_the_detailed_scale_back(authed):
    """Тот самый выход: день недоступен на всём проекте, но доступен на окне."""
    project_id, _, _ = _make_project(
        authed, tasks=20, duration=60, start=date.today() - timedelta(days=30)
    )
    response = authed.get(
        f"/api/projects/{project_id}/export.pdf?include=gantt&zoom=day&period=next_4w"
    )
    assert response.status_code == 200
    assert _pdf_pages(response.content) == 1


def test_a_relative_plan_has_no_today_and_says_so(authed):
    """У плана без дат ось — «День N», и окно «ближайшие 4 недели» на ней не
    определено. Это отсутствие величины, а не отказ по вкусу."""
    project_id = authed.post("/api/projects", json={"name": "Черновик"}).json()["id"]
    category_id = _mutate(
        authed, project_id, {"type": "create_category", "name": "Этап", "color": "#3b82f6"}
    )["op"]["category_id"]
    _mutate(
        authed,
        project_id,
        {
            "type": "create_task",
            "category_id": category_id,
            "name": "Задача",
            "start_date": "2001-01-01",
            "duration_days": 4,
        },
    )

    refused = authed.get(
        f"/api/projects/{project_id}/export.pdf?include=gantt&period=next_4w"
    )
    assert refused.status_code == 422
    assert refused.json()["detail"] == "export_period_undated"

    # «Весь проект» на той же оси работает: отказ касается только окон,
    # отсчитываемых от сегодня.
    assert authed.get(
        f"/api/projects/{project_id}/export.pdf?include=gantt&period=all"
    ).status_code == 200


def test_portrait_needs_more_pages_than_landscape_at_the_same_scale():
    """Ёмкость страницы считается формулой от ширины, а не таблицей чисел."""
    days = 365
    for zoom in budget.ZOOMS:
        assert budget.page_count(days, zoom, Orientation.PORTRAIT) >= budget.page_count(
            days, zoom, Orientation.LANDSCAPE
        )
    assert budget.page_count(days, Zoom.DAY, Orientation.PORTRAIT) > budget.page_count(
        days, Zoom.DAY, Orientation.LANDSCAPE
    )


def test_the_default_zoom_is_never_forbidden_to_itself():
    """На десятилетнем портфеле даже месяц выходит за потолок — и он всё равно
    разрешён: у него нет менее подробного соседа, и отказ там означал бы, что
    проект не выгружается вовсе, а это уже не защита от неподъёмного файла."""
    for days in (30, 365, 1095, 5000, 20000):
        zoom = budget.default_zoom(days, Orientation.LANDSCAPE)
        assert budget.allowed(zoom, days, Orientation.LANDSCAPE)

    assert budget.page_count(5000, Zoom.MONTH, Orientation.LANDSCAPE) > budget.MAX_PAGES
    assert budget.allowed(Zoom.MONTH, 5000, Orientation.LANDSCAPE)


def test_a_decade_long_project_still_exports(authed):
    """Обратная сторона: правило проверяется не только арифметикой, но и
    маршрутом — иначе потолок однажды переехал бы в него отдельной строкой."""
    project_id, _, _ = _make_project(authed, tasks=30, duration=180)
    response = authed.get(f"/api/projects/{project_id}/export.pdf?include=gantt")
    assert response.status_code == 200, response.text


def test_the_gantt_sheet_never_outgrows_the_column_ceiling():
    """Лист ленты на страницы не режется — он одна широкая полоса, и предел
    ему ставится по числу колонок."""
    for days in (30, 365, 1095, 5000):
        zoom = budget.default_zoom_for_xlsx(days)
        assert budget.columns_for(days, zoom) <= budget.MAX_XLSX_COLUMNS


def test_the_window_of_a_period_never_leaves_the_project(authed):
    whole = budget.Window(date(2026, 1, 1), date(2026, 12, 31))
    today = date(2026, 6, 1)
    for period in Period:
        window = budget.resolve_window(period, whole, today, dated=True)
        assert whole.start <= window.start <= window.end <= whole.end


def test_a_project_entirely_in_the_past_still_gets_a_window(authed):
    """Пустая шкала — это страница, на которой ничего нет; лучше показать
    конец проекта, чем пустоту."""
    whole = budget.Window(date(2024, 1, 1), date(2024, 3, 1))
    window = budget.resolve_window(Period.NEXT_4W, whole, date(2026, 6, 1), dated=True)
    assert window.start == window.end == whole.end


# --- скоркард: метрики с разной историей ----------------------------------------
#
# Метрики появляются и снимаются (см. миграцию scorecard_signal_cleanup), а
# снимки прошлых недель неизменяемы: у метрики, заведённой в августе, июльских
# снимков нет и не будет. Значит набор недель у метрик разный, и таблица обязана
# это переживать.


def _metric(key: str, history: list[tuple[str, float]], value=0.0, status="ok") -> dict:
    return {
        "key": key,
        "value": value,
        "status": status,
        "history": [
            {"week_start": week, "value": point, "status": "ok"} for week, point in history
        ],
    }


def test_the_week_columns_are_the_union_across_metrics():
    """Иначе метрика помоложе обрезала бы историю всем остальным."""
    old = _metric("overdue_tasks", [("2026-08-03", 1), ("2026-08-10", 2)])
    young = _metric("finish_drift", [("2026-08-10", 3)])

    weeks = align_weeks([old, young], date(2026, 8, 17))

    assert weeks == [date(2026, 8, 3), date(2026, 8, 10), date(2026, 8, 17)]


def test_a_metric_without_a_snapshot_gets_a_gap_not_a_shift():
    """Прочерк на своём месте, а не сдвиг соседних значений влево: сдвинутая
    строка — это молча неверный документ, и заметили бы его не сразу."""
    old = _metric("overdue_tasks", [("2026-08-03", 1), ("2026-08-10", 2)], value=5)
    young = _metric("finish_drift", [("2026-08-10", 3)], value=7)
    weeks = align_weeks([old, young], date(2026, 8, 17))

    assert metric_row(old, weeks) == ([1, 2, 5], ["ok", "ok", "ok"])
    # У молодой метрики первая колонка пуста, а её единственный снимок стоит
    # под своей неделей — второй, а не первой.
    values, statuses = metric_row(young, weeks)
    assert values == [None, 3, 7]
    assert statuses == ["no_data", "ok", "ok"]


def test_a_first_metric_without_history_does_not_collapse_the_table():
    """Тот самый случай, который сегодня не наступает лишь из-за порядка метрик
    в миграции: раньше недели брались у первой метрики, и молодая на нулевой
    позиции схлопнула бы таблицу в одну колонку."""
    young = _metric("finish_drift", [])
    old = _metric("overdue_tasks", [("2026-08-03", 1), ("2026-08-10", 2)])

    weeks = align_weeks([young, old], date(2026, 8, 17))

    assert len(weeks) == 3
    assert metric_row(young, weeks) == ([None, None, 0.0], ["no_data", "no_data", "ok"])


def test_the_current_week_is_always_the_last_column():
    """Её значение живое и берётся не из истории — снимка за неё может ещё не
    быть вовсе."""
    metric = _metric("overdue_tasks", [("2026-08-03", 1)], value=9, status="risk")
    weeks = align_weeks([metric], date(2026, 8, 17))

    assert weeks[-1] == date(2026, 8, 17)
    assert metric_row(metric, weeks) == ([1, 9], ["ok", "risk"])


def test_a_snapshot_for_the_current_week_does_not_double_the_column():
    """Снимок текущей недели уже есть (ленивая фиксация при открытии
    скоркарда) — колонка всё равно одна, и в ней живое значение."""
    metric = _metric("overdue_tasks", [("2026-08-17", 4)], value=6)
    weeks = align_weeks([metric], date(2026, 8, 17))

    assert weeks == [date(2026, 8, 17)]
    assert metric_row(metric, weeks) == ([6], ["ok"])


def test_an_unknown_metric_status_does_not_cost_the_whole_document():
    """Набор состояний живёт в ScorecardStatus и меняется вместе со скоркардом.
    Падение на неизвестном — пятисотка на весь файл из-за одной ячейки."""
    assert theme.metric_cell("risk") == theme.METRIC_CELL["risk"]
    assert theme.metric_cell("brand_new") == theme.METRIC_CELL["no_data"]


# --- словари -------------------------------------------------------------------


def test_every_language_has_every_label():
    """Полнота словарей проверяется здесь, а не глазами: рассинхрон копится
    незаметно и обнаруживается уже выгруженным файлом."""
    locales = available_locales()
    assert set(locales) >= {"az", "en", "ru"}

    reference = {(group, key) for group, keys in dictionary("az").items() for key in keys}
    for locale in locales:
        actual = {
            (group, key) for group, keys in dictionary(locale).items() for key in keys
        }
        assert actual == reference, f"словарь {locale} разошёлся: {reference ^ actual}"


def test_every_mutation_has_a_name_in_the_history_dictionary():
    """Журнал правок подписывает операции по имени. Новая операция без подписи
    не должна ронять документ — но и молча превращаться в «Правка плана» ей
    незачем, пока о ней помнят здесь."""
    import re

    source = Path(__file__).resolve().parents[1] / "app" / "mutations.py"
    declared = set(re.findall(r'type: Literal\["([a-z_]+)"\]', source.read_text()))
    assert declared, "не удалось вычитать список операций из mutations.py"

    named = set(dictionary("ru")["event"]) - {"unknown"}
    assert declared <= named, f"нет подписи для операций: {sorted(declared - named)}"


def test_the_embedded_font_covers_all_three_languages():
    """Встроенные шрифты ReportLab — Latin-1: без вшитого Inter документ на
    двух языках из трёх остался бы без букв."""
    from reportlab.pdfbase.ttfonts import TTFont

    fonts = Path(__file__).resolve().parents[1] / "app" / "export" / "fonts"
    for file in ("Inter-Regular.ttf", "Inter-SemiBold.ttf", "Inter-Bold.ttf"):
        face = TTFont("probe", str(fonts / file)).face
        missing = [ch for ch in "АБВЯабвяəğşıİçöüÇÖÜ0123456789№◆·—" if ord(ch) not in face.charToGlyph]
        assert missing == [], f"{file} не покрывает {missing}"
